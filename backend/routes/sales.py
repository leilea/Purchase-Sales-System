from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, SalesOrder, SalesOrderItem, SalesOrderSupplier, SalesOrderLogistics, PurchaseOrder, PurchaseOrderItem, Inventory, InventoryLog, Product, Customer, Supplier, Unit
from datetime import datetime
import uuid

bp = Blueprint('sales', __name__)


def serialize_order(o):
    return {
        'id': o.id,
        'order_no': o.order_no,
        'customer_id': o.customer_id,
        'customer_name': o.customer.name if o.customer else None,
        'contact_person': o.contact_person,
        'contact_phone': o.contact_phone,
        'business_manager': o.business_manager,
        'business_manager_phone': o.business_manager_phone,
        'order_date': o.order_date.isoformat() if o.order_date else None,
        'address': o.address,
        'freight_responsible': o.freight_responsible,
        'freight': float(o.freight) if o.freight else 0,
        'payment_method': o.payment_method,
        'invoice_required': o.invoice_required or 0,
        'invoice_tax_rate': o.invoice_tax_rate,
        'total_amount': float(o.total_amount) if o.total_amount else 0,
        'gross_profit': float(o.gross_profit) if o.gross_profit else 0,
        'status': o.status,
        'remark': o.remark,
        'created_by': o.created_by,
        'created_at': o.created_at.isoformat() if o.created_at else None,
        'items': [serialize_item(i) for i in o.items],
        'suppliers': [serialize_supplier(s) for s in o.suppliers],
        'logistics': [serialize_logistics(l) for l in o.logistics]
    }


def serialize_supplier(s):
    return {
        'id': s.id,
        'supplier': s.supplier,
        'product': s.product,
        'spec': s.spec,
        'grade': s.grade,
        'surface_treatment': s.surface_treatment,
        'quantity': float(s.quantity) if s.quantity else 0,
        'total_weight': float(s.total_weight) if s.total_weight else 0,
        'unit_weight': float(s.unit_weight) if s.unit_weight else 0,
        'price': float(s.price) if s.price else 0,
        'packaging': s.packaging,
        'tax': float(s.tax) if s.tax else 0
    }


def serialize_logistics(l):
    return {
        'id': l.id,
        'company': l.company,
        'contact_person': l.contact_person,
        'contact_phone': l.contact_phone,
        'start_time': l.start_time.isoformat() if l.start_time else None,
        'end_time': l.end_time.isoformat() if l.end_time else None,
        'transport_status': l.transport_status,
        'freight': float(l.freight) if l.freight else 0,
        'forklift_fee': float(l.forklift_fee) if l.forklift_fee else 0,
        'other_fee': float(l.other_fee) if l.other_fee else 0,
        'remark': l.remark
    }


def serialize_item(i):
    return {
        'id': i.id,
        'product_id': i.product_id,
        'product_name': i.product.name if i.product else None,
        'product_code': i.product.code if i.product else None,
        'spec': i.spec,
        'material_grade': i.material_grade,
        'surface_treatment': i.surface_treatment,
        'matching': i.matching,
        'unit_name': i.product.unit.name if i.product and i.product.unit else None,
        'quantity': float(i.quantity) if i.quantity else 0,
        'unit_price': float(i.unit_price) if i.unit_price else 0,
        'amount': float(i.amount) if i.amount else 0,
        'remark': i.remark
    }


@bp.route('', methods=['GET'])
@jwt_required()
def list_sales():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    customer_id = request.args.get('customer_id', type=int)
    order_no = request.args.get('order_no', '')
    customer_name = request.args.get('customer_name', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    query = SalesOrder.query
    if status:
        query = query.filter_by(status=status)
    if customer_id:
        query = query.filter_by(customer_id=customer_id)
    if order_no:
        query = query.filter(SalesOrder.order_no.like(f'%{order_no}%'))
    if customer_name:
        query = query.join(Customer).filter(Customer.name.like(f'%{customer_name}%'))
    if start_date:
        query = query.filter(SalesOrder.order_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(SalesOrder.order_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    pagination = query.order_by(SalesOrder.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'items': [serialize_order(o) for o in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'page': page
    })


@bp.route('/<int:id>', methods=['GET'])
@jwt_required()
def get_sale(id):
    order = SalesOrder.query.get_or_404(id)
    return jsonify(serialize_order(order))


@bp.route('', methods=['POST'])
@jwt_required()
def create_sale():
    user_id = int(get_jwt_identity())
    data = request.get_json()
    
    order_no = f"SO{datetime.now().strftime('%Y%m%d')}{uuid.uuid4().hex[:6].upper()}"
    
    order = SalesOrder(
        order_no=order_no,
        customer_id=data.get('customer_id'),
        contact_person=data.get('contact_person'),
        contact_phone=data.get('contact_phone'),
        business_manager=data.get('business_manager'),
        business_manager_phone=data.get('business_manager_phone'),
        address=data.get('address'),
        freight_responsible=data.get('freight_responsible'),
        freight=data.get('freight', 0),
        payment_method=data.get('payment_method'),
        invoice_required=data.get('invoice_required', 0),
        invoice_tax_rate=data.get('invoice_tax_rate'),
        order_date=datetime.strptime(data.get('order_date', datetime.now().strftime('%Y-%m-%d')), '%Y-%m-%d').date() if data.get('order_date') else datetime.now().date(),
        total_amount=0,
        gross_profit=data.get('gross_profit', 0),
        status='draft',
        remark=data.get('remark'),
        created_by=user_id
    )
    db.session.add(order)
    db.session.flush()
    
    items_total = 0
    for item_data in data.get('items', []):
        quantity = float(item_data.get('quantity', 0))
        unit_price = float(item_data.get('unit_price', 0))
        amount = quantity * unit_price
        
        item = SalesOrderItem(
            order_id=order.id,
            product_id=item_data.get('product_id'),
            spec=item_data.get('spec'),
            material_grade=item_data.get('material_grade'),
            surface_treatment=item_data.get('surface_treatment'),
            matching=item_data.get('matching'),
            quantity=quantity,
            unit_price=unit_price,
            amount=amount,
            remark=item_data.get('remark')
        )
        db.session.add(item)
        items_total += amount
    
    order.total_amount = items_total + float(data.get('freight', 0))
    
    for sup_data in data.get('suppliers', []):
        supplier_rec = SalesOrderSupplier(
            order_id=order.id,
            supplier=sup_data.get('supplier'),
            product=sup_data.get('product'),
            spec=sup_data.get('spec'),
            grade=sup_data.get('grade'),
            surface_treatment=sup_data.get('surface_treatment'),
            quantity=sup_data.get('quantity', 0),
            total_weight=sup_data.get('total_weight', 0),
            unit_weight=sup_data.get('unit_weight', 0),
            price=sup_data.get('price', 0),
            packaging=sup_data.get('packaging'),
            tax=sup_data.get('tax', 0)
        )
        db.session.add(supplier_rec)
    
    for log_data in data.get('logistics', []):
        log_rec = SalesOrderLogistics(
            order_id=order.id,
            company=log_data.get('company'),
            contact_person=log_data.get('contact_person'),
            contact_phone=log_data.get('contact_phone'),
            start_time=datetime.fromisoformat(log_data['start_time']) if log_data.get('start_time') else None,
            end_time=datetime.fromisoformat(log_data['end_time']) if log_data.get('end_time') else None,
            transport_status=log_data.get('transport_status'),
            freight=log_data.get('freight', 0),
            forklift_fee=log_data.get('forklift_fee', 0),
            other_fee=log_data.get('other_fee', 0),
            remark=log_data.get('remark')
        )
        db.session.add(log_rec)
    
    db.session.commit()
    
    return jsonify(serialize_order(order)), 201


@bp.route('/<int:id>', methods=['PUT'])
@jwt_required()
def update_sale(id):
    order = SalesOrder.query.get_or_404(id)
    if order.status != 'draft':
        return jsonify({'error': 'Only draft orders can be updated'}), 400
    
    data = request.get_json()
    
    if 'customer_id' in data:
        order.customer_id = data['customer_id']
    if 'contact_person' in data:
        order.contact_person = data['contact_person']
    if 'contact_phone' in data:
        order.contact_phone = data['contact_phone']
    if 'business_manager' in data:
        order.business_manager = data['business_manager']
    if 'business_manager_phone' in data:
        order.business_manager_phone = data['business_manager_phone']
    if 'address' in data:
        order.address = data['address']
    if 'freight_responsible' in data:
        order.freight_responsible = data['freight_responsible']
    if 'freight' in data:
        order.freight = data['freight']
    if 'payment_method' in data:
        order.payment_method = data['payment_method']
    if 'invoice_required' in data:
        order.invoice_required = data['invoice_required']
    if 'invoice_tax_rate' in data:
        order.invoice_tax_rate = data['invoice_tax_rate']
    if 'order_date' in data:
        order.order_date = datetime.strptime(data['order_date'], '%Y-%m-%d').date()
    if 'remark' in data:
        order.remark = data['remark']
    if 'gross_profit' in data:
        order.gross_profit = data['gross_profit']
    
    if 'items' in data:
        SalesOrderItem.query.filter_by(order_id=order.id).delete()
        items_total = 0
        for item_data in data['items']:
            quantity = float(item_data.get('quantity', 0))
            unit_price = float(item_data.get('unit_price', 0))
            amount = quantity * unit_price
            
            item = SalesOrderItem(
                order_id=order.id,
                product_id=item_data.get('product_id'),
                spec=item_data.get('spec'),
                material_grade=item_data.get('material_grade'),
                surface_treatment=item_data.get('surface_treatment'),
                matching=item_data.get('matching'),
                quantity=quantity,
                unit_price=unit_price,
                amount=amount,
                remark=item_data.get('remark')
            )
            db.session.add(item)
            items_total += amount
        order.total_amount = items_total + float(data.get('freight', order.freight or 0))
    
    if 'suppliers' in data:
        SalesOrderSupplier.query.filter_by(order_id=order.id).delete()
        for sup_data in data['suppliers']:
            supplier_rec = SalesOrderSupplier(
                order_id=order.id,
                supplier=sup_data.get('supplier'),
                product=sup_data.get('product'),
                spec=sup_data.get('spec'),
                grade=sup_data.get('grade'),
                surface_treatment=sup_data.get('surface_treatment'),
                quantity=sup_data.get('quantity', 0),
                total_weight=sup_data.get('total_weight', 0),
                unit_weight=sup_data.get('unit_weight', 0),
                price=sup_data.get('price', 0),
                packaging=sup_data.get('packaging'),
                tax=sup_data.get('tax', 0)
            )
            db.session.add(supplier_rec)
    
    if 'logistics' in data:
        SalesOrderLogistics.query.filter_by(order_id=order.id).delete()
        for log_data in data['logistics']:
            log_rec = SalesOrderLogistics(
                order_id=order.id,
                company=log_data.get('company'),
                contact_person=log_data.get('contact_person'),
                contact_phone=log_data.get('contact_phone'),
                start_time=datetime.fromisoformat(log_data['start_time']) if log_data.get('start_time') else None,
                end_time=datetime.fromisoformat(log_data['end_time']) if log_data.get('end_time') else None,
                transport_status=log_data.get('transport_status'),
                freight=log_data.get('freight', 0),
                forklift_fee=log_data.get('forklift_fee', 0),
                other_fee=log_data.get('other_fee', 0),
                remark=log_data.get('remark')
            )
            db.session.add(log_rec)
    
    db.session.commit()
    return jsonify(serialize_order(order))


@bp.route('/upload', methods=['POST'])
@jwt_required()
def upload_sales():
    user_id = int(get_jwt_identity())

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']

    import openpyxl
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    def cell_str(row, col):
        val = ws.cell(row, col).value
        if val is None:
            return ''
        if isinstance(val, float):
            if val == int(val):
                return str(int(val))
            return f'{val:.2f}'
        return str(val).strip()

    def extract_after_prefix(text, *prefixes):
        for p in prefixes:
            if p in text:
                return text.split(p, 1)[1].strip()
        return text

    def float_safe(val):
        if val is None:
            return 0
        s = str(val).strip()
        if s.startswith('='):
            return 0
        try:
            return float(s.replace(',', ''))
        except (ValueError, TypeError):
            return 0

    raw1 = cell_str(3, 1)
    customer_name = extract_after_prefix(raw1, '：', ':')
    raw_addr = cell_str(3, 5)
    address = extract_after_prefix(raw_addr, '：', ':')
    contact_person = extract_after_prefix(cell_str(4, 1), '：', ':')
    phone = extract_after_prefix(cell_str(4, 5), '：', ':')
    payment_method = extract_after_prefix(cell_str(5, 1), '：', ':')

    raw_invoice = cell_str(5, 5)
    invoice_tax_rate = ''
    if '：' in raw_invoice:
        val = raw_invoice.split('：', 1)[1].strip()
        if '13' in val: invoice_tax_rate = '13%'
        elif '9' in val: invoice_tax_rate = '9%'
        elif '6' in val: invoice_tax_rate = '6%'
        elif '3' in val: invoice_tax_rate = '3%'
        elif '0' in val: invoice_tax_rate = '0%'
    invoice_required = 1 if ('13' in raw_invoice or '开' in raw_invoice or '是' in raw_invoice) else 0

    raw_freight = cell_str(5, 10)
    freight_responsible = '己方承担'
    if '：' in raw_freight or ':' in str(ws.cell(5, 10).value or ''):
        val = raw_freight.split('：', 1)[1].strip() if '：' in raw_freight else str(ws.cell(5, 10).value).split(':', 1)[1].strip()
        if '客户' in val:
            freight_responsible = '客户承担'

    biz_manager = ''
    if cell_str(6, 1) and '：' in cell_str(6, 1):
        biz_manager = cell_str(6, 1).split('：', 1)[1].strip()

    remark_header = ''
    if cell_str(7, 1) and '：' in cell_str(7, 1):
        remark_header = cell_str(7, 1).split('：', 1)[1].strip()

    date_val = ws.cell(4, 10).value
    if isinstance(date_val, datetime):
        order_date = date_val.strftime('%Y-%m-%d')
    elif isinstance(date_val, float) and date_val > 40000:
        from datetime import timedelta
        order_date = (datetime(1899, 12, 30) + timedelta(days=int(date_val))).strftime('%Y-%m-%d')
    else:
        order_date = datetime.now().strftime('%Y-%m-%d')

    customer = None
    if customer_name:
        customer = Customer.query.filter(Customer.name.like(f'%{customer_name}%')).first()
    if not customer:
        customer = Customer(
            code=f'IMP{datetime.now().strftime("%Y%m%d%H%M%S")}',
            name=customer_name or '导入客户',
            contact=contact_person,
            phone=phone or '',
            address=address or ''
        )
        db.session.add(customer)
        db.session.flush()

    col_map = {
        'name': 2, 'spec': 3, 'material_grade': 4, 'surface_treatment': 5,
        'matching': 8, 'unit': 9, 'quantity': 10, 'unit_price': 11, 'remark': 12
    }

    order = SalesOrder(
        order_no=f"SO{datetime.now().strftime('%Y%m%d')}{uuid.uuid4().hex[:6].upper()}",
        customer_id=customer.id,
        contact_person=contact_person,
        contact_phone=phone or '',
        address=address or '',
        business_manager=biz_manager,
        freight_responsible=freight_responsible,
        payment_method=payment_method,
        invoice_required=invoice_required,
        invoice_tax_rate=invoice_tax_rate,
        freight=0,
        order_date=datetime.strptime(order_date, '%Y-%m-%d').date(),
        total_amount=0,
        gross_profit=0,
        status='draft',
        remark=remark_header or '',
        created_by=user_id
    )
    db.session.add(order)
    db.session.flush()

    items_total = 0
    for row in range(10, ws.max_row + 1):
        seq_raw = ws.cell(row, 1).value
        if seq_raw is None:
            continue
        seq_str = str(seq_raw).strip()
        if not seq_str:
            continue
        try:
            int(seq_str)
        except (ValueError, TypeError):
            if '合计' in seq_str or 'sum' in seq_str.lower():
                break
            if '供应商' in seq_str or '委外' in seq_str:
                break
            continue

        product_name = cell_str(row, col_map['name'])
        if not product_name:
            continue

        spec = cell_str(row, col_map['spec'])
        mg = cell_str(row, col_map['material_grade'])
        st = cell_str(row, col_map['surface_treatment'])
        matching = cell_str(row, col_map['matching'])
        unit_name = cell_str(row, col_map['unit'])

        qty_s = cell_str(row, col_map['quantity'])
        price_s = cell_str(row, col_map['unit_price'])
        item_remark_raw = cell_str(row, col_map['remark'])
        item_remark = '' if item_remark_raw.startswith('=') else item_remark_raw

        try:
            quantity = float(qty_s.replace(',', ''))
        except (ValueError, TypeError):
            quantity = 0
        try:
            unit_price = float(price_s.replace(',', ''))
        except (ValueError, TypeError):
            unit_price = 0
        amount = quantity * unit_price

        product = Product.query.filter(Product.name.like(f'%{product_name}%')).first()
        if not product:
            unit_obj = None
            if unit_name:
                unit_obj = Unit.query.filter(Unit.name == unit_name).first()
                if not unit_obj:
                    unit_obj = Unit(name=unit_name)
                    db.session.add(unit_obj)
                    db.session.flush()

            product = Product(
                code=f'IMP{datetime.now().strftime("%Y%m%d%H%M%S")}{uuid.uuid4().hex[:4].upper()}',
                name=product_name,
                spec=spec,
                material_grade=mg,
                surface_treatment=st,
                unit_id=unit_obj.id if unit_obj else None,
                cost_price=0,
                sale_price=unit_price,
                status=1
            )
            db.session.add(product)
            db.session.flush()

        item = SalesOrderItem(
            order_id=order.id,
            product_id=product.id,
            spec=spec or product.spec,
            material_grade=mg or product.material_grade,
            surface_treatment=st or product.surface_treatment,
            matching=matching,
            quantity=quantity,
            unit_price=unit_price,
            amount=amount,
            remark=item_remark
        )
        db.session.add(item)
        items_total += amount

    order.total_amount = items_total

    # Parse supplier/outsourcing rows
    supplier_start = None
    for r in range(12, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and '供应商' in str(val):
            supplier_start = r + 2
            break

    if supplier_start:
        for row in range(supplier_start, ws.max_row + 1):
            seq_raw = ws.cell(row, 1).value
            if seq_raw is None:
                continue
            seq_str = str(seq_raw).strip()
            if not seq_str:
                continue
            try:
                int(seq_str)
            except (ValueError, TypeError):
                if '合计' in seq_str:
                    break
                continue

            supplier_name = cell_str(row, 2)
            sup = SalesOrderSupplier(
                order_id=order.id,
                supplier=supplier_name,
                product=cell_str(row, 3),
                spec=cell_str(row, 4),
                grade=cell_str(row, 5),
                surface_treatment=cell_str(row, 6),
                quantity=float_safe(ws.cell(row, 7).value),
                unit_weight=float_safe(ws.cell(row, 9).value),
                price=float_safe(ws.cell(row, 10).value),
                packaging=cell_str(row, 11),
                tax=float_safe(ws.cell(row, 12).value)
            )
            db.session.add(sup)

    # Parse logistics rows
    logistics_start = None
    for r in range(14, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and ('物流' in str(val) or '运输' in str(val)):
            logistics_start = r + 2
            break

    if logistics_start:
        for row in range(logistics_start, ws.max_row + 1):
            seq_raw = ws.cell(row, 1).value
            if seq_raw is None:
                break
            seq_str = str(seq_raw).strip()
            if not seq_str:
                break
            try:
                int(seq_str)
            except (ValueError, TypeError):
                break

            start_val = ws.cell(row, 5).value
            end_val = ws.cell(row, 6).value

            log_rec = SalesOrderLogistics(
                order_id=order.id,
                company=cell_str(row, 2),
                contact_person=cell_str(row, 3),
                contact_phone=cell_str(row, 4),
                start_time=start_val if isinstance(start_val, datetime) else None,
                end_time=end_val if isinstance(end_val, datetime) else None,
                transport_status=cell_str(row, 7),
                freight=float_safe(ws.cell(row, 10).value),
                forklift_fee=float_safe(ws.cell(row, 11).value),
                remark=cell_str(row, 12)
            )
            db.session.add(log_rec)

    supplier_cost = sum(
        float(s.price or 0) * float(s.quantity or 0) + float(s.tax or 0)
        for s in order.suppliers.all()
    )
    logistics_cost = sum(
        float(l.freight or 0) + float(l.forklift_fee or 0) + float(l.other_fee or 0)
        for l in order.logistics.all()
    )
    order.gross_profit = order.total_amount - supplier_cost - logistics_cost

    db.session.commit()

    return jsonify(serialize_order(order)), 201


@bp.route('/<int:id>/deliver', methods=['POST'])
@jwt_required()
def deliver_sale(id):
    user_id = int(get_jwt_identity())
    order = SalesOrder.query.get_or_404(id)
    
    if order.status == 'delivered':
        return jsonify({'error': 'Order already delivered'}), 400
    
    if order.status == 'cancelled':
        return jsonify({'error': 'Cancelled order cannot be delivered'}), 400
    
    for item in order.items:
        product = Product.query.get(item.product_id)
        if not product:
            continue
        
        inventory = Inventory.query.filter_by(product_id=item.product_id).first()
        if not inventory or inventory.quantity < item.quantity:
            continue
        
        inventory.quantity -= item.quantity
        
        log = InventoryLog(
            product_id=item.product_id,
            change_type='sales_out',
            quantity=-item.quantity,
            order_no=order.order_no
        )
        db.session.add(log)
    
    if order.customer and not Customer.query.filter(Customer.name == order.customer.name).first():
        code = f'IMP{datetime.now().strftime("%Y%m%d%H%M%S")}{uuid.uuid4().hex[:4].upper()}'
        db.session.add(Customer(
            code=code,
            name=order.customer.name,
            contact=order.contact_person or '',
            phone=order.contact_phone or '',
            address=order.address or ''
        ))

    seen_suppliers = set()
    for sup in order.suppliers:
        name = sup.supplier
        if name and name not in seen_suppliers:
            seen_suppliers.add(name)
            if not Supplier.query.filter(Supplier.name == name).first():
                code = f'IMP{datetime.now().strftime("%Y%m%d%H%M%S")}{uuid.uuid4().hex[:4].upper()}'
                db.session.add(Supplier(name=name, code=code, contact='', phone=''))

    db.session.flush()

    po_index = 0
    for sup in order.suppliers:
        name = sup.supplier
        if not name:
            continue

        supplier = Supplier.query.filter(Supplier.name == name).first()
        if not supplier:
            continue

        po_index += 1
        po_no = f'{order.order_no}-{po_index}'

        po = PurchaseOrder(
            order_no=po_no,
            supplier_id=supplier.id,
            order_date=datetime.now().date(),
            total_amount=0,
            status='received',
            remark=f'由销售订单 {order.order_no} 自动生成',
            created_by=user_id
        )
        db.session.add(po)
        db.session.flush()

        total_amount = 0
        product = None
        if sup.product:
            product = Product.query.filter(Product.name.like(f'%{sup.product}%')).first()
        if not product and sup.product:
            product = Product(
                code=f'IMP{datetime.now().strftime("%Y%m%d%H%M%S")}{uuid.uuid4().hex[:4].upper()}',
                name=sup.product,
                spec=sup.spec or '',
                material_grade=sup.grade or '',
                surface_treatment=sup.surface_treatment or '',
                cost_price=float(sup.price) if sup.price else 0,
                sale_price=0,
                status=1
            )
            db.session.add(product)
            db.session.flush()
        if product:
            quantity = float(sup.quantity) if sup.quantity else 0
            unit_price = float(sup.price) if sup.price else 0
            amount = quantity * unit_price

            po_item = PurchaseOrderItem(
                order_id=po.id,
                product_id=product.id,
                quantity=quantity,
                unit_price=unit_price,
                amount=amount
            )
            db.session.add(po_item)
            total_amount = amount

        po.total_amount = total_amount

    order.status = 'delivered'
    db.session.commit()
    
    return jsonify(serialize_order(order))


@bp.route('/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_sale(id):
    order = SalesOrder.query.get_or_404(id)
    if order.status not in ('draft', 'cancelled'):
        return jsonify({'error': 'Only draft or cancelled orders can be deleted'}), 400
    db.session.delete(order)
    db.session.commit()
    return jsonify({'message': 'Order deleted successfully'})


@bp.route('/<int:id>/cancel', methods=['POST'])
@jwt_required()
def cancel_sale(id):
    order = SalesOrder.query.get_or_404(id)
    
    if order.status == 'delivered':
        return jsonify({'error': 'Delivered order cannot be cancelled'}), 400
    
    if order.status == 'cancelled':
        return jsonify({'error': 'Order already cancelled'}), 400
    
    order.status = 'cancelled'
    db.session.commit()
    
    return jsonify(serialize_order(order))
